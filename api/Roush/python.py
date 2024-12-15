import random
from flask_cors import CORS
from flask import Flask, jsonify  
from openpyxl import load_workbook

# Initialize Flask app
app = Flask(__name__)
CORS(app)  # Enable CORS for cross-origin requests

@app.route('/api/questions', methods=['GET'])
def get_questions():
    # Load data from Excel workbook
    workbook = load_workbook("book1.xlsx")
    sheet_1 = workbook["Sheet1"]
    sheet_2 = workbook["Sheet2"]

    # Extract data from sheets
    data_two_marks = [list(row) for row in sheet_1.iter_rows(values_only=True)]
    headers_two_marks = data_two_marks[0]
    questions_two_marks = data_two_marks[1:]

    data_ten_marks = [list(row) for row in sheet_2.iter_rows(values_only=True)]
    headers_ten_marks = data_ten_marks[0]
    questions_ten_marks = data_ten_marks[1:]

    # Try to find the "Bloom's Level" column in both sheets
    try:
        bloom_col_two = headers_two_marks.index("Bloom's Level")
        bloom_col_ten = headers_ten_marks.index("Bloom's Level")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400  # Return error if Bloom's Level column is missing

    # Filter rows where Bloom's Level is not None
    filtered_two_marks = [row for row in questions_two_marks if row[bloom_col_two] is not None]
    filtered_ten_marks = [row for row in questions_ten_marks if row[bloom_col_ten] is not None]

    # Group questions by Bloom's Level
    def group_by_bloom_level(questions, bloom_col):
        bloom_groups = {}
        for question in questions:
            level = question[bloom_col]
            if level not in bloom_groups:
                bloom_groups[level] = []
            bloom_groups[level].append(question)
        return bloom_groups

    two_marks_groups = group_by_bloom_level(filtered_two_marks, bloom_col_two)
    ten_marks_groups = group_by_bloom_level(filtered_ten_marks, bloom_col_ten)

    # Select questions randomly from the groups
    def select_questions(grouped_questions, total_questions=5):
        selected_questions = []
        all_levels = list(grouped_questions.keys())
        random.shuffle(all_levels)  # Shuffle levels for randomness
        while len(selected_questions) < total_questions and all_levels:
            for level in all_levels:
                if grouped_questions[level]:
                    selected_questions.append(grouped_questions[level].pop())
                if len(selected_questions) >= total_questions:
                    break
            all_levels = [level for level in all_levels if grouped_questions[level]]
        return selected_questions

    selected_two_marks = select_questions(two_marks_groups, 5)
    selected_ten_marks = select_questions(ten_marks_groups, 5)

    # Prepare data for response (S.No, Question, Bloom's Level)
    two_marks_data = [
        [idx + 1, question[1], question[bloom_col_two]]  # Include S.No, Question, Bloom's Level
        for idx, question in enumerate(selected_two_marks)
    ]
    ten_marks_data = [
        [idx + 1, question[1], question[bloom_col_ten]]  # Include S.No, Question, Bloom's Level
        for idx, question in enumerate(selected_ten_marks)
    ]

    # Return the selected questions as a JSON response
    return jsonify({
        "two_marks": {
            "questions": two_marks_data
        },
        "ten_marks": {
            "questions": ten_marks_data
        }
    })

if __name__ == '__main__':
    # Run the Flask app on port 5001
    app.run(debug=True, port=5001)
