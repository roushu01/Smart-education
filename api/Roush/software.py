import random
from flask import Flask, jsonify  
from flask_cors import CORS 
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)  

@app.route('/api/questions', methods=['GET'])
def get_questions():
   
    workbook = load_workbook("software.xlsx")
    sheet_10 = workbook["ten_marks"]
    sheet_2 = workbook["two_marks"]
    data_two_marks = [list(row) for row in sheet_2.iter_rows(values_only=True)]
    headers_two_marks = data_two_marks[0]
    questions_two_marks = data_two_marks[1:]

    data_ten_marks = [list(row) for row in sheet_10.iter_rows(values_only=True)]
    headers_ten_marks = data_ten_marks[0]
    questions_ten_marks = data_ten_marks[1:]

    # Identify Bloom's Level column
    try:
        bloom_col_two = headers_two_marks.index("Bloom's Level")
        bloom_col_ten = headers_ten_marks.index("Bloom's Level")
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    # Filter rows with valid "Bloom's Level" values
    filtered_two_marks = [row for row in questions_two_marks if row[bloom_col_two] is not None]
    filtered_ten_marks = [row for row in questions_ten_marks if row[bloom_col_ten] is not None]

    # Group questions by Bloom's Level
    def group_by_bloom_level(questions, bloom_col):
        bloom_groups = {}
        for question in questions:
            level = question[bloom_col]
            if level not in bloom_groups:
                bloom_groups[level] = []
            bloom_groups[level].append(question)
        return bloom_groups

    two_marks_groups = group_by_bloom_level(filtered_two_marks, bloom_col_two)
    ten_marks_groups = group_by_bloom_level(filtered_ten_marks, bloom_col_ten)

    # Select questions
    def select_questions(grouped_questions, total_questions=5):
        selected_questions = []
        all_levels = list(grouped_questions.keys())
        random.shuffle(all_levels)
        while len(selected_questions) < total_questions and all_levels:
            for level in all_levels:
                if grouped_questions[level]:
                    selected_questions.append(grouped_questions[level].pop())
                if len(selected_questions) >= total_questions:
                    break
            all_levels = [level for level in all_levels if grouped_questions[level]]
        return selected_questions

    selected_two_marks = select_questions(two_marks_groups, 5)
    selected_ten_marks = select_questions(ten_marks_groups, 5)

    # Add serial numbers
    selected_two_marks = [[idx + 1] + question for idx, question in enumerate(selected_two_marks)]
    selected_ten_marks = [[idx + 1] + question for idx, question in enumerate(selected_ten_marks)]

    # Add headers for S.No
    headers_two_marks = ["S.No"] + headers_two_marks
    headers_ten_marks = ["S.No"] + headers_ten_marks

    return jsonify({
        "two_marks": {
            "headers": headers_two_marks,
            "questions": selected_two_marks
        },
        "ten_marks": {
            "headers": headers_ten_marks,
            "questions": selected_ten_marks
        }
    })

if __name__ == '__main__':
    app.run(debug=True)
